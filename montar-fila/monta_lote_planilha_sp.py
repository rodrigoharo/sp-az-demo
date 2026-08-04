#!/usr/bin/env python
from __future__ import annotations

import argparse
import base64
import csv
import json
import re
import shutil
import subprocess
import sys
import warnings
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from shared_code.dam_normalization import normalize_dam_folder_segment, normalize_dam_name_component, strip_accents


GRAPH_ROOT = "https://graph.microsoft.com/v1.0"
SHAREPOINT_HOST = "https://contoso.sharepoint.com/"
warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")


@dataclass(frozen=True)
class ResolvedFolder:
    hostname: str
    site_path: str
    site_name: str
    drive_id: str
    folder_path: str


@dataclass(frozen=True)
class SourceEntry:
    key: str
    folder: str
    metadata_value: str
    source_url: str
    subfolder: str = ""


FIELD_MAP = {
    "referenciaColecaoLote": "B",
    "ciclo": "D",
    "cdISA": "E",
    "cdRede": "F",
    "anoUso": "G",
    "segmento": "H",
    "selo": "I",
    "colecao": "J",
    "nome": "K",
    "produto": "L",
    "saida": "M",
    "statusProduto": "N",
    "qtdPaginas": "O",
    "pathSharePointArqAbertos": "P",
    "pathSharePointPDFFinal": "Q",
    "obs": "R",
    "retrancaAnterior": "S",
    "sintaxeNova": "T",
    "tagsFase1": "U",
}

IDENTIFIER_FIELDS = {"cdISA", "cdRede"}
INTEGER_FIELDS = {"qtdPaginas"}

SOURCE_COLUMNS = {
    "arquivos-abertos": {
        "column": "P",
        "folder": "Arquivos Abertos",
        "metadata_value": "Arquivos Abertos",
    },
    "pdf-final": {
        "column": "Q",
        "folder": "PDF - FINAL",
        "metadata_value": "PDF - FINAL",
    },
}

URL_IN_TEXT_RE = re.compile(r"https?://[^\s<>\"]+", re.IGNORECASE)
SOURCE_LABEL_RE = re.compile(r"^\s*(capa|capas|miolo)(?:\b|[\s:_-]|$)", re.IGNORECASE)

DAM_CATEGORY_RULES = [
    ("grafica", ("impresso", "grafica", "grafico")),
    ("digital", ("digital", "ebook", "e-book", "objeto digital")),
    ("pnld", ("pnld",)),
    ("acessibilidade", ("acessibilidade", "accessibility")),
    ("Politicas-e-Procedimentos", ("politica", "politicas", "procedimento", "procedimentos")),
    ("Manuais-e-Guias", ("manual", "manuais", "guia", "guias")),
    ("Contratos-e-Garantias", ("contrato", "contratos", "garantia", "garantias")),
    ("documentacao-geral", ("documentacao", "documento", "documentos")),
]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Monta mensagens de fila a partir da planilha de ativos DAM."
    )
    parser.add_argument("--workbook", required=True, help="Caminho do arquivo .xlsx.")
    parser.add_argument("--sheet", default="ATIVOS-LT-1-2", help="Nome da aba.")
    parser.add_argument("--start-row", type=int, default=5, help="Primeira linha valida.")
    parser.add_argument("--end-row", type=int, help="Ultima linha da planilha a considerar.")
    parser.add_argument("--row", type=int, action="append", help="Processa apenas esta linha. Pode repetir.")
    parser.add_argument("--max-rows", type=int, help="Limita a quantidade de linhas lidas.")
    parser.add_argument(
        "--filter-tag",
        action="append",
        help="Processa somente linhas cuja coluna TAGS - FASE 1 contenha esta tag. Pode repetir.",
    )
    parser.add_argument("--target-blob-root", default="sp/ativos", help="Raiz do caminho no Blob.")
    parser.add_argument("--output-dir", required=True, help="Pasta de saida.")
    parser.add_argument(
        "--graph-auth",
        choices=["function-app", "env", "azure-cli"],
        default="function-app",
        help="Origem da autenticacao Graph. Padrao: function-app.",
    )
    parser.add_argument("--resource-group", default="rg-demo-dam", help="Resource group da Function.")
    parser.add_argument(
        "--function-app-name",
        default="sp-dam-migration-worker",
        help="Nome da Function App usada para ler SHAREPOINT_*.",
    )
    parser.add_argument(
        "--no-individual-files",
        action="store_true",
        help="Nao gera um .json individual por mensagem.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    token = get_graph_token(args)
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Aba nao encontrada: {args.sheet}. Disponiveis: {', '.join(workbook.sheetnames)}")

    ws = workbook[args.sheet]
    if args.end_row is not None and args.end_row < args.start_row:
        raise SystemExit("--end-row deve ser maior ou igual a --start-row.")

    rows_to_process = set(args.row or [])
    last_row = min(args.end_row or ws.max_row, ws.max_row)
    messages: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    candidate_rows = collect_candidate_rows(
        ws,
        args.start_row,
        last_row,
        rows_to_process,
        args.max_rows,
        args.filter_tag or [],
    )
    message_index_by_key: dict[tuple[str, str, str, str, str, str, str], int] = {}

    for row_number, row_metadata in candidate_rows:
        dam_path = build_dam_path(
            args.target_blob_root,
            row_metadata,
            row_number,
        )

        for source_key, source_cfg in SOURCE_COLUMNS.items():
            cell = ws[f"{source_cfg['column']}{row_number}"]
            for source_entry in iter_source_entries(cell, source_key, source_cfg):
                source_url = source_entry.source_url
                if not source_url:
                    continue

                try:
                    resolved = resolve_sharepoint_folder_url(source_url, token)
                    target_blob_prefix = clean_queue_path(
                        "/".join(
                            [
                                dam_path["prefix"],
                                normalize_path_segment(source_entry.folder),
                                normalize_path_segment(source_entry.subfolder) if source_entry.subfolder else "",
                            ]
                        )
                    )
                    source_label = source_entry_label(source_entry)

                    metadata = build_blob_metadata(
                        row_number=row_number,
                        source_type=source_entry.metadata_value,
                        resolved=resolved,
                        row_metadata=row_metadata,
                        source_url=source_url,
                        dam_path=dam_path,
                        source_subfolder=source_entry.subfolder,
                    )
                    aem_metadata = build_aem_metadata(
                        row_metadata,
                        source_url,
                        source_entry.metadata_value,
                        source_entry.subfolder,
                    )
                    aem_metadata.update(build_aem_dam_metadata(dam_path, source_entry.metadata_value))

                    message_key = (
                        source_entry.key,
                        resolved.hostname,
                        resolved.site_path,
                        resolved.drive_id,
                        resolved.folder_path,
                        target_blob_prefix,
                        source_entry.subfolder,
                    )
                    existing_index = message_index_by_key.get(message_key)
                    if existing_index is not None:
                        existing_message = messages[existing_index]
                        merge_metadata(existing_message["metadata"], metadata)
                        merge_metadata(existing_message["aemMetadata"], aem_metadata)
                        merge_summary(summary[existing_index], row_number, row_metadata)
                        continue

                    message = {
                        "sharePointHostname": resolved.hostname,
                        "sharePointSitePath": resolved.site_path,
                        "sharePointDriveId": resolved.drive_id,
                        "sharePointFolderPath": resolved.folder_path,
                        "targetBlobPrefix": target_blob_prefix,
                        "recursive": True,
                        "metadata": metadata,
                        "aemMetadata": aem_metadata,
                    }
                    message_index_by_key[message_key] = len(messages)
                    messages.append(message)
                    summary.append(
                        {
                            "linha": row_number,
                            "linhasAgrupadas": str(row_number),
                            "origem": source_entry.metadata_value,
                            "subpastaOrigem": source_entry.subfolder,
                            "codigoISA": row_metadata.get("cdISA") or "",
                            "ciclo": row_metadata.get("ciclo") or "",
                            "codigoColecao": dam_path["codigoColecao"],
                            "colecao": row_metadata.get("colecao") or "",
                            "nome": row_metadata.get("nome") or "",
                            "retrancaMaterial": dam_path["retrancaMaterial"],
                            "categoriaDestino": dam_path["categoriaLabel"],
                            "tipoDestino": source_label,
                            "regraAplicada": dam_path["rule"],
                            "sharePointFolderPath": resolved.folder_path,
                            "targetBlobPrefix": target_blob_prefix,
                            "jsonFile": "",
                        }
                    )
                except Exception as exc:  # noqa: BLE001
                    errors.append(
                        {
                            "linha": row_number,
                            "origem": source_entry_label(source_entry),
                            "url": source_url,
                            "erro": str(exc),
                        }
                    )

    jsonl_path = output_dir / "mensagens-fila.jsonl"
    with jsonl_path.open("w", encoding="utf-8", newline="\n") as fp:
        for message in messages:
            fp.write(json.dumps(message, ensure_ascii=False) + "\n")

    if not args.no_individual_files:
        for idx, message in enumerate(messages, start=1):
            linha = message["metadata"]["linhaPlanilha"]
            origem = slugify(message["metadata"]["origemCaminho"])
            file_name = f"linha-{first_line_number(linha):04d}-{origem}-{idx:04d}.json"
            file_path = output_dir / file_name
            file_path.write_text(
                json.dumps(message, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            summary[idx - 1]["jsonFile"] = file_name

    write_csv(output_dir / "mensagens-fila-resumo.csv", summary)
    write_csv(output_dir / "mensagens-fila-erros.csv", errors)

    print("CONCLUIDO")
    print(f"Workbook: {workbook_path}")
    print(f"Aba: {args.sheet}")
    print(f"Linhas validas lidas: {len(candidate_rows)}")
    print(f"Mensagens geradas: {len(messages)}")
    print(f"Erros: {len(errors)}")
    print(f"JSONL: {jsonl_path}")
    print(f"Resumo: {output_dir / 'mensagens-fila-resumo.csv'}")
    if errors:
        print(f"Erros: {output_dir / 'mensagens-fila-erros.csv'}")
    return 0 if not errors else 2


def get_graph_token(args: argparse.Namespace) -> str:
    if args.graph_auth == "function-app":
        settings = get_function_app_settings(args.resource_group, args.function_app_name)
        return get_client_credentials_token(
            settings.get("SHAREPOINT_TENANT_ID", ""),
            settings.get("SHAREPOINT_CLIENT_ID", ""),
            settings.get("SHAREPOINT_CLIENT_SECRET", ""),
        )

    if args.graph_auth == "env":
        return get_client_credentials_token(
            read_env("SHAREPOINT_TENANT_ID"),
            read_env("SHAREPOINT_CLIENT_ID"),
            read_env("SHAREPOINT_CLIENT_SECRET"),
        )

    return get_azure_cli_graph_token()


def get_azure_cli_graph_token() -> str:
    az_cli = find_azure_cli()
    command = [
        az_cli,
        "account",
        "get-access-token",
        "--resource-type",
        "ms-graph",
        "--query",
        "accessToken",
        "-o",
        "tsv",
    ]
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
    except FileNotFoundError as exc:
        raise RuntimeError("Azure CLI nao encontrado. Faca login com az login antes de rodar.") from exc
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"Falha ao obter token do Graph via Azure CLI: {exc.stderr.strip()}") from exc

    token = result.stdout.strip()
    if not token:
        raise RuntimeError("Azure CLI retornou token vazio para o Graph.")
    return token


def get_function_app_settings(resource_group: str, function_app_name: str) -> dict[str, str]:
    az_cli = find_azure_cli()
    command = [
        az_cli,
        "functionapp",
        "config",
        "appsettings",
        "list",
        "-g",
        resource_group,
        "-n",
        function_app_name,
        "-o",
        "json",
    ]
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            "Falha ao ler app settings da Function. "
            f"Use --graph-auth azure-cli ou confira permissoes. {exc.stderr.strip()}"
        ) from exc

    values = json.loads(result.stdout)
    return {item["name"]: item.get("value", "") for item in values}


def get_client_credentials_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    if not tenant_id or not client_id or not client_secret:
        raise RuntimeError("SHAREPOINT_TENANT_ID, SHAREPOINT_CLIENT_ID e SHAREPOINT_CLIENT_SECRET sao obrigatorios.")

    body = urlencode(
        {
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        }
    ).encode("utf-8")
    req = Request(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urlopen(req, timeout=60) as response:
        payload = json.loads(response.read().decode("utf-8"))
    token = payload.get("access_token")
    if not token:
        raise RuntimeError("Token vazio retornado pelo Microsoft identity platform.")
    return token


def read_env(name: str) -> str:
    import os

    return os.environ.get(name, "")


def find_azure_cli() -> str:
    for candidate in ("az", "az.cmd", "az.exe"):
        path = shutil.which(candidate)
        if path:
            return path

    windows_default = Path(r"C:\Program Files\Microsoft SDKs\Azure\CLI2\wbin\az.cmd")
    if windows_default.exists():
        return str(windows_default)

    return "az"


def graph_get(url: str, token: str) -> dict[str, Any]:
    req = Request(url, headers={"Authorization": f"Bearer {token}"})
    with urlopen(req, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def resolve_sharepoint_folder_url(url: str, token: str) -> ResolvedFolder:
    parsed = urlparse(url)
    hostname = parsed.netloc
    if not hostname:
        raise ValueError(f"URL sem host: {url}")

    if re.search(r"/:[A-Za-z]:/", url):
        sharing_token = sharing_url_token(url)
        item = graph_get(f"{GRAPH_ROOT}/shares/{sharing_token}/driveItem", token)
        parent = item.get("parentReference") or {}
        drive_id = parent.get("driveId")
        if not drive_id:
            raise ValueError("Graph nao retornou parentReference.driveId.")

        drive_path = re.sub(r"^/drives/[^/]+/root:", "", parent.get("path") or "").strip("/")
        item_name = item.get("name")
        if item_name:
            drive_path = clean_queue_path(f"{drive_path}/{item_name}")

        site = graph_get(f"{GRAPH_ROOT}/sites/{parent['siteId']}", token)
        site_path = urlparse(site["webUrl"]).path.rstrip("/")
        return ResolvedFolder(
            hostname=urlparse(site["webUrl"]).netloc or hostname,
            site_path=site_path,
            site_name=site_path.rstrip("/").rsplit("/", 1)[-1],
            drive_id=drive_id,
            folder_path=drive_path,
        )

    query = parse_qs(parsed.query)
    server_rel_path = unquote((query.get("id") or [""])[0]).lstrip("/")
    if not server_rel_path:
        raise ValueError("URL sem parametro id= e sem formato de compartilhamento :f:/.")

    parts = server_rel_path.split("/")
    if len(parts) < 3 or parts[0] not in {"sites", "teams", "personal"}:
        raise ValueError(f"Nao foi possivel interpretar o parametro id=: /{server_rel_path}")

    site_path = f"/{parts[0]}/{parts[1]}"
    library_name = parts[2]
    folder_path = clean_queue_path("/".join(parts[3:]))

    site = graph_get(f"{GRAPH_ROOT}/sites/{hostname}:{site_path}", token)
    drives = graph_get(f"{GRAPH_ROOT}/sites/{site['id']}/drives", token)
    drive = first_matching_drive(drives.get("value") or [], library_name)
    if not drive:
        available = ", ".join(d.get("name", "") for d in drives.get("value") or [])
        raise ValueError(f"Biblioteca '{library_name}' nao encontrada. Disponiveis: {available}")

    return ResolvedFolder(
        hostname=hostname,
        site_path=site_path,
        site_name=parts[1],
        drive_id=drive["id"],
        folder_path=folder_path,
    )


def first_matching_drive(drives: list[dict[str, Any]], library_name: str) -> dict[str, Any] | None:
    decoded_library_name = unquote(library_name)
    for drive in drives:
        drive_name = drive.get("name") or ""
        web_url = unquote(drive.get("webUrl") or "")
        if drive_name == decoded_library_name or web_url.endswith(f"/{decoded_library_name}"):
            return drive
    return None


def sharing_url_token(url: str) -> str:
    raw = base64.urlsafe_b64encode(url.encode("utf-8")).decode("ascii").rstrip("=")
    return f"u!{raw}"


def get_cell_link(cell: Any) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    value = cell.value
    if isinstance(value, str) and value.strip().lower().startswith(("http://", "https://")):
        return value.strip()
    return ""


def normalize_sharepoint_url(target: str) -> str:
    if not target:
        return ""
    normalized = target.strip().replace("\\", "/")
    sp_token = r":[A-Za-z]:/(?:r/sites|s)/.*"

    match = re.match(rf"^(?:\.\./)+({sp_token})$", normalized)
    if match:
        return SHAREPOINT_HOST + match.group(1)

    match = re.search(rf"/(?:\.\./)+({sp_token})$", normalized)
    if match:
        return SHAREPOINT_HOST + match.group(1)

    if re.match(rf"^{sp_token}$", normalized):
        return SHAREPOINT_HOST + normalized

    match = re.match(
        r"^(?:\.\./)+([^/]+)/(Documentos(?:%20|\s)(?:Compartilhados|Partilhados)/.*)$",
        normalized,
        re.IGNORECASE,
    )
    if match:
        site, rest = match.groups()
        return f"{SHAREPOINT_HOST}sites/{site}/{rest}"

    return normalized


def iter_source_entries(cell: Any, source_key: str, source_cfg: dict[str, str]) -> list[SourceEntry]:
    entries: list[tuple[str, str]] = []
    hyperlink_url = normalize_sharepoint_url(get_cell_link(cell))
    hyperlink_label = first_source_label(cell.value)
    if hyperlink_url:
        entries.append((hyperlink_label, hyperlink_url))

    entries.extend(parse_labeled_source_urls(cell.value))

    if not entries:
        return []

    result: list[SourceEntry] = []
    seen: set[tuple[str, str]] = set()
    for subfolder, source_url in entries:
        normalized_url = normalize_sharepoint_url(source_url)
        if not normalized_url:
            continue
        dedupe_key = (subfolder, normalized_url)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        key_suffix = f":{subfolder}" if subfolder else ""
        result.append(
            SourceEntry(
                key=f"{source_key}{key_suffix}",
                folder=source_cfg["folder"],
                metadata_value=source_cfg["metadata_value"],
                source_url=normalized_url,
                subfolder=subfolder,
            )
        )
    return result


def parse_labeled_source_urls(value: Any) -> list[tuple[str, str]]:
    text = clean_value(value)
    if not text:
        return []

    current_label = ""
    entries: list[tuple[str, str]] = []
    for raw_line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line:
            continue

        label = first_source_label(line)
        if label:
            current_label = label

        for match in URL_IN_TEXT_RE.findall(line):
            source_url = clean_extracted_url(match)
            entries.append((first_source_label(line) or source_label_from_url(source_url) or current_label, source_url))

    if not entries and text.lower().startswith(("http://", "https://")):
        entries.append(("", clean_extracted_url(text)))

    return entries


def source_label_from_url(value: str) -> str:
    text = strip_accents(unquote(clean_value(value))).lower()
    parts = re.split(r"[/\\?#&=]+", text)
    for part in parts:
        if SOURCE_LABEL_RE.match(part):
            return first_source_label(part)
    return ""


def first_source_label(value: Any) -> str:
    text = clean_value(value)
    if not text:
        return ""
    first_line = next((line.strip() for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if line.strip()), "")
    match = SOURCE_LABEL_RE.match(first_line)
    if not match:
        return ""
    raw_label = strip_accents(match.group(1)).lower()
    if raw_label.startswith("capa"):
        return "capa"
    if raw_label.startswith("miolo"):
        return "miolo"
    return ""


def clean_extracted_url(value: str) -> str:
    return value.strip().rstrip(").,;]")


def source_entry_label(source_entry: SourceEntry) -> str:
    if not source_entry.subfolder:
        return source_entry.folder
    return f"{source_entry.folder}/{source_entry.subfolder}"


def read_row_metadata(ws: Any, row_number: int) -> dict[str, str]:
    values: dict[str, str] = {}
    for field, column in FIELD_MAP.items():
        cell = ws[f"{column}{row_number}"]
        if column in {"P", "Q"}:
            raw_value = clean_value(cell.value, field)
            value = raw_value if parse_labeled_source_urls(raw_value) else normalize_sharepoint_url(get_cell_link(cell)) or raw_value
        else:
            value = clean_value(cell.value, field)
        values[field] = value
    return values


def is_valid_row(metadata: dict[str, str]) -> bool:
    return bool(metadata.get("colecao") and metadata.get("nome"))


def collect_candidate_rows(
    ws: Any,
    start_row: int,
    last_row: int,
    rows_to_process: set[int],
    max_rows: int | None,
    filter_tags: list[str],
) -> list[tuple[int, dict[str, str]]]:
    rows: list[tuple[int, dict[str, str]]] = []
    for row_number in range(start_row, last_row + 1):
        if rows_to_process and row_number not in rows_to_process:
            continue
        if max_rows is not None and len(rows) >= max_rows:
            break

        row_metadata = read_row_metadata(ws, row_number)
        if not is_valid_row(row_metadata):
            continue
        if not row_matches_filter_tags(row_metadata, filter_tags):
            continue
        rows.append((row_number, row_metadata))
    return rows


def row_matches_filter_tags(metadata: dict[str, str], filter_tags: list[str]) -> bool:
    if not filter_tags:
        return True

    row_tags = set(split_tags(metadata.get("tagsFase1", "")))
    if not row_tags:
        return False

    wanted_tags = {normalize_tag(tag) for tag in filter_tags if normalize_tag(tag)}
    return bool(row_tags.intersection(wanted_tags))


def split_tags(value: str) -> list[str]:
    return [
        normalize_tag(part)
        for part in re.split(r"[,;|\n]+", str(value))
        if normalize_tag(part)
    ]


def normalize_tag(value: str) -> str:
    return strip_accents(str(value)).strip().upper()


def merge_metadata(target: dict[str, str], source: dict[str, str]) -> None:
    for key, value in source.items():
        value_text = str(value).strip()
        if not value_text:
            continue
        current = str(target.get(key, "")).strip()
        if not current:
            target[key] = value_text
            continue
        if current == value_text:
            continue
        if value_text not in split_merged_values(current):
            target[key] = f"{current} | {value_text}"


def merge_summary(target: dict[str, Any], row_number: int, row_metadata: dict[str, str]) -> None:
    target["linhasAgrupadas"] = append_unique(
        str(target.get("linhasAgrupadas", "")),
        str(row_number),
        separator=";",
    )
    target["linha"] = first_line_number(target.get("linhasAgrupadas", target.get("linha", row_number)))
    for summary_key, metadata_key in (
        ("codigoISA", "cdISA"),
        ("colecao", "colecao"),
        ("nome", "nome"),
    ):
        value = row_metadata.get(metadata_key) or ""
        if value:
            target[summary_key] = append_unique(str(target.get(summary_key, "")), value)


def append_unique(current: str, value: str, separator: str = " | ") -> str:
    value = str(value).strip()
    if not value:
        return current
    if current.strip() == value:
        return current
    values = split_merged_values(current, separator)
    if value not in values:
        values.append(value)
    return separator.join(values)


def split_merged_values(value: str, separator: str = " | ") -> list[str]:
    return [part.strip() for part in str(value).split(separator) if part.strip()]


def first_line_number(value: object) -> int:
    match = re.search(r"\d+", str(value))
    return int(match.group(0)) if match else 0


def build_blob_metadata(
    *,
    row_number: int,
    source_type: str,
    resolved: ResolvedFolder,
    row_metadata: dict[str, str],
    source_url: str,
    dam_path: dict[str, str],
    source_subfolder: str = "",
) -> dict[str, str]:
    metadata = {
        "origem": "sharepoint",
        "origemCaminho": source_type,
        "linhaPlanilha": str(row_number),
        "site": resolved.site_name,
        "sharePointFolderPath": resolved.folder_path,
        "sharePointSourceUrl": source_url,
        "damRoot": dam_path["root"],
        "damCiclo": dam_path["ciclo"],
        "damTipo": dam_path["tipo"],
        "damCodigoColecao": dam_path["codigoColecao"],
        "damRetrancaMaterial": dam_path["retrancaMaterial"],
        "damCategoria": dam_path["categoriaLabel"],
    }
    if source_subfolder:
        metadata["origemSubpasta"] = source_subfolder
    return metadata


def build_aem_metadata(
    row_metadata: dict[str, str],
    source_url: str,
    source_type: str,
    source_subfolder: str = "",
) -> dict[str, str]:
    metadata = {f"demo:{key}": value for key, value in row_metadata.items() if value}
    metadata["demo:origemCaminho"] = source_type
    metadata["demo:sharePointSourceUrl"] = source_url
    if source_subfolder:
        metadata["demo:origemSubpasta"] = source_subfolder
    return metadata


def build_aem_dam_metadata(dam_path: dict[str, str], source_type: str) -> dict[str, str]:
    return {
        "demo:damRoot": dam_path["root"],
        "demo:damCiclo": dam_path["ciclo"],
        "demo:damTipo": dam_path["tipo"],
        "demo:damCodigoColecao": dam_path["codigoColecao"],
        "demo:damRetrancaMaterial": dam_path["retrancaMaterial"],
        "demo:damCategoria": dam_path["categoriaLabel"],
        "demo:damOrigemConteudo": source_type,
        "demo:damRegraTaxonomia": dam_path["rule"],
    }


def build_dam_path(
    target_blob_root: str,
    row_metadata: dict[str, str],
    row_number: int,
) -> dict[str, str]:
    root = clean_queue_path(target_blob_root or "sp/ativos")
    ciclo = normalize_path_segment(row_metadata.get("ciclo") or "sem-ciclo")
    codigo_colecao = normalize_path_segment(
        row_metadata.get("referenciaColecaoLote")
        or row_metadata.get("cdRede")
        or row_metadata.get("colecao")
        or "sem-colecao"
    )
    item_label = row_metadata.get("sintaxeNova") or join_non_empty([row_metadata.get("cdISA"), row_metadata.get("nome")], " - ")
    retranca_material = normalize_retranca_segment(item_label or f"Linha {row_number}")
    tipo = "backup"
    categoria_label, categoria_rule = resolve_dam_category(row_metadata)
    categoria = normalize_path_segment(categoria_label)
    prefix = clean_queue_path("/".join([root, ciclo, tipo, codigo_colecao, retranca_material, categoria]))
    return {
        "root": root,
        "ciclo": ciclo,
        "tipo": tipo,
        "codigoColecao": codigo_colecao,
        "retrancaMaterial": retranca_material,
        "categoria": categoria,
        "categoriaLabel": categoria_label,
        "prefix": prefix,
        "rule": f"sp/ativos/{{ciclo}}/backup/{{codigoColecao}}/{{retrancaMaterial}}/{{categoria}}/{{origem}}; {categoria_rule}",
    }


def resolve_dam_category(row_metadata: dict[str, str]) -> tuple[str, str]:
    source_value = row_metadata.get("saida") or ""
    normalized_source = normalize_for_rule_match(source_value)
    for category, keywords in DAM_CATEGORY_RULES:
        if any(keyword in normalized_source for keyword in keywords):
            return category, f"categoria por SAIDA={source_value or 'vazio'}"
    return "documentacao-geral", f"categoria padrao; SAIDA={source_value or 'vazio'}"


def normalize_for_rule_match(value: str) -> str:
    text = strip_accents(clean_value(value)).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def clean_value(value: Any, field: str | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if field in IDENTIFIER_FIELDS:
        return normalize_identifier(text)
    if field in INTEGER_FIELDS:
        return normalize_integer(text)
    return text


def normalize_identifier(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    candidate = text.replace(",", ".")
    if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", candidate):
        return text
    if "e" not in candidate.lower():
        return text[:-2] if text.endswith(".0") else text
    try:
        number = Decimal(candidate)
    except InvalidOperation:
        return text
    if number == number.to_integral_value():
        return format(number.quantize(Decimal(1)), "f")
    return format(number.normalize(), "f").rstrip("0").rstrip(".")


def normalize_integer(value: str) -> str:
    text = value.strip()
    candidate = text.replace(",", ".")
    if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", candidate):
        return text
    try:
        number = Decimal(candidate)
    except InvalidOperation:
        return text
    if number == number.to_integral_value():
        return format(number.quantize(Decimal(1)), "f")
    return text


def normalize_path_segment(value: str) -> str:
    return normalize_dam_folder_segment(clean_value(value))


def normalize_retranca_segment(value: str) -> str:
    return normalize_dam_name_component(clean_value(value))


def clean_queue_path(path: str) -> str:
    return re.sub(r"/+", "/", path.strip("/"))


def join_non_empty(values: list[str], separator: str) -> str:
    return separator.join(str(value).strip() for value in values if str(value).strip())


def compare_retranca_key(value: str) -> str:
    return " ".join(clean_value(value).upper().split())


def slugify(value: str) -> str:
    text = strip_accents(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "mensagem"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames: list[str]
    if rows:
        fieldnames = list(rows[0].keys())
    else:
        fieldnames = ["linha", "origem", "erro"]
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    sys.exit(main())

