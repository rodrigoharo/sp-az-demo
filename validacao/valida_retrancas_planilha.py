#!/usr/bin/env python
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from shared_code.dam_normalization import normalize_dam_name_component


FIELD_MAP = {
    "colecao": "J",
    "nome": "K",
    "pathSharePointArqAbertos": "P",
    "pathSharePointPDFFinal": "Q",
    "retrancaAnterior": "S",
    "sintaxeNova": "T",
}

REPORT_COLUMNS = [
    "severidade",
    "tipo",
    "linha",
    "retrancaAnterior",
    "sintaxeNova",
    "sintaxeNovaNormalizada",
    "detalhe",
]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Valida retranca anterior e sintaxe nova antes de enviar lote para a fila."
    )
    parser.add_argument("--workbook", required=True, help="Caminho da planilha .xlsx.")
    parser.add_argument("--sheet", default="ATIVOS-LT-1-2", help="Nome da aba.")
    parser.add_argument("--start-row", type=int, default=5, help="Primeira linha valida.")
    parser.add_argument("--end-row", type=int, help="Ultima linha da planilha a considerar.")
    parser.add_argument("--output-dir", required=True, help="Pasta para gravar relatorios.")
    parser.add_argument(
        "--fail-on-error",
        action="store_true",
        help="Retorna exit code 2 quando houver erro bloqueante.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Aba nao encontrada: {args.sheet}. Disponiveis: {', '.join(workbook.sheetnames)}")

    ws = workbook[args.sheet]
    if args.end_row is not None and args.end_row < args.start_row:
        raise SystemExit("--end-row deve ser maior ou igual a --start-row.")

    last_row = min(args.end_row or ws.max_row, ws.max_row)
    rows = [read_row(ws, row_number) for row_number in range(args.start_row, last_row + 1)]
    rows = [row for row in rows if is_candidate_row(row)]

    issues = validate_rows(rows)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"validacao_retrancas_{timestamp}.csv"
    json_path = output_dir / f"validacao_retrancas_{timestamp}.json"
    html_path = output_dir / f"validacao_retrancas_{timestamp}.html"

    write_csv(csv_path, issues)
    write_json(json_path, workbook_path, args.sheet, args.start_row, last_row, rows, issues)
    write_html(html_path, workbook_path, args.sheet, args.start_row, last_row, rows, issues)

    errors = [issue for issue in issues if issue["severidade"] == "ERRO"]
    warnings = [issue for issue in issues if issue["severidade"] == "ALERTA"]

    print("CONCLUIDO")
    print(f"Linhas avaliadas: {len(rows)}")
    print(f"Erros: {len(errors)}")
    print(f"Alertas: {len(warnings)}")
    print(f"CSV: {csv_path}")
    print(f"JSON: {json_path}")
    print(f"HTML: {html_path}")

    if errors:
        print()
        print("Primeiros erros:")
        for issue in errors[:10]:
            print(f"- linha {issue['linha']}: {issue['tipo']} - {issue['detalhe']}")

    return 2 if args.fail_on_error and errors else 0


def read_row(ws: Any, row_number: int) -> dict[str, str]:
    row = {"linha": str(row_number)}
    for field, column in FIELD_MAP.items():
        cell = ws[f"{column}{row_number}"]
        row[field] = clean_value(cell.value)
        if column in {"P", "Q"} and cell.hyperlink and cell.hyperlink.target:
            row[field] = clean_value(cell.hyperlink.target)
    row["sintaxeNovaNormalizada"] = normalize_retranca(row["sintaxeNova"])
    return row


def is_candidate_row(row: dict[str, str]) -> bool:
    return any(
        row.get(field)
        for field in (
            "colecao",
            "nome",
            "pathSharePointArqAbertos",
            "pathSharePointPDFFinal",
            "retrancaAnterior",
            "sintaxeNova",
        )
    )


def validate_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []

    for row in rows:
        old = row["retrancaAnterior"]
        new = row["sintaxeNova"]
        has_source = bool(row["pathSharePointArqAbertos"] or row["pathSharePointPDFFinal"])

        if not old:
            issues.append(issue("ERRO", "RETRANCA_ANTERIOR_VAZIA", row, "Coluna S vazia em linha candidata."))
        if not new:
            issues.append(issue("ERRO", "SINTAXE_NOVA_VAZIA", row, "Coluna T vazia em linha candidata."))
        if not has_source:
            issues.append(issue("ALERTA", "SEM_LINK_ORIGEM", row, "Linha sem link nas colunas P e Q."))

    old_to_new: dict[str, list[dict[str, str]]] = defaultdict(list)
    new_to_old: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        old_key = normalize_compare_key(row["retrancaAnterior"])
        new_key = row["sintaxeNovaNormalizada"]
        if old_key and new_key:
            old_to_new[old_key].append(row)
            new_to_old[new_key].append(row)

    for grouped_rows in old_to_new.values():
        distinct_new = sorted({row["sintaxeNovaNormalizada"] for row in grouped_rows if row["sintaxeNovaNormalizada"]})
        if len(distinct_new) <= 1:
            continue
        if is_la_lp_pair(distinct_new):
            continue
        rows_text = ", ".join(row["linha"] for row in grouped_rows)
        detail = f"Mesma retranca anterior aponta para sintaxes novas diferentes: {', '.join(distinct_new)}. Linhas: {rows_text}."
        for row in grouped_rows:
            issues.append(issue("ALERTA", "MESMA_RETRANCA_ANTERIOR_COM_SINTAXES_DIFERENTES", row, detail))

    for new_key, grouped_rows in new_to_old.items():
        distinct_old = sorted({normalize_compare_key(row["retrancaAnterior"]) for row in grouped_rows if row["retrancaAnterior"]})
        if len(distinct_old) <= 1:
            continue
        rows_text = ", ".join(row["linha"] for row in grouped_rows)
        detail = f"Sintaxe nova normalizada '{new_key}' aparece para retrancas anteriores diferentes. Linhas: {rows_text}."
        for row in grouped_rows:
            issues.append(issue("ERRO", "SINTAXE_NOVA_DUPLICADA_PARA_RETRANCAS_DIFERENTES", row, detail))

    return sorted(issues, key=lambda item: (severity_rank(item["severidade"]), int(item["linha"]), item["tipo"]))


def issue(severity: str, issue_type: str, row: dict[str, str], detail: str) -> dict[str, str]:
    return {
        "severidade": severity,
        "tipo": issue_type,
        "linha": row["linha"],
        "retrancaAnterior": row.get("retrancaAnterior", ""),
        "sintaxeNova": row.get("sintaxeNova", ""),
        "sintaxeNovaNormalizada": row.get("sintaxeNovaNormalizada", ""),
        "detalhe": detail,
    }


def severity_rank(severity: str) -> int:
    return {"ERRO": 0, "ALERTA": 1}.get(severity, 9)


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_retranca(value: str) -> str:
    text = clean_value(value)
    return normalize_dam_name_component(text) if text else ""


def normalize_compare_key(value: str) -> str:
    return " ".join(clean_value(value).upper().split())


def is_la_lp_pair(variants: list[str]) -> bool:
    bases: set[str] = set()
    suffixes: set[str] = set()
    for variant in variants:
        match = re.match(r"^(?P<base>.+)-(?P<suffix>la|lp)$", variant)
        if not match:
            return False
        bases.add(match.group("base"))
        suffixes.add(match.group("suffix"))
    return len(bases) == 1 and suffixes == {"la", "lp"}


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=REPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_json(
    path: Path,
    workbook_path: Path,
    sheet: str,
    start_row: int,
    end_row: int,
    rows: list[dict[str, str]],
    issues: list[dict[str, str]],
) -> None:
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(workbook_path),
        "sheet": sheet,
        "startRow": start_row,
        "endRow": end_row,
        "rowsEvaluated": len(rows),
        "errors": sum(1 for issue in issues if issue["severidade"] == "ERRO"),
        "warnings": sum(1 for issue in issues if issue["severidade"] == "ALERTA"),
        "issues": issues,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_html(
    path: Path,
    workbook_path: Path,
    sheet: str,
    start_row: int,
    end_row: int,
    rows: list[dict[str, str]],
    issues: list[dict[str, str]],
) -> None:
    errors = sum(1 for issue in issues if issue["severidade"] == "ERRO")
    warnings = sum(1 for issue in issues if issue["severidade"] == "ALERTA")
    table_rows = "\n".join(
        "<tr>"
        + "".join(f"<td>{html.escape(str(issue.get(column, '')))}</td>" for column in REPORT_COLUMNS)
        + "</tr>"
        for issue in issues
    )
    if not table_rows:
        table_rows = f"<tr><td colspan='{len(REPORT_COLUMNS)}'>Nenhum problema encontrado.</td></tr>"
    header = "".join(f"<th>{html.escape(column)}</th>" for column in REPORT_COLUMNS)
    path.write_text(
        f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <title>Validacao de retrancas</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 24px; color: #111827; }}
    .summary {{ background: #eef6ff; border: 1px solid #bfdbfe; padding: 12px; margin-bottom: 18px; }}
    .cards {{ display: flex; gap: 12px; margin-bottom: 18px; }}
    .card {{ border: 1px solid #d1d5db; border-radius: 6px; padding: 12px; min-width: 160px; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    th, td {{ border: 1px solid #d1d5db; padding: 6px 8px; text-align: left; vertical-align: top; }}
    th {{ background: #f3f4f6; }}
  </style>
</head>
<body>
  <h1>Validacao de retrancas</h1>
  <div class="summary">
    <div><strong>Workbook:</strong> {html.escape(str(workbook_path))}</div>
    <div><strong>Aba:</strong> {html.escape(sheet)}</div>
    <div><strong>Intervalo:</strong> {start_row} a {end_row}</div>
  </div>
  <div class="cards">
    <div class="card"><strong>Linhas avaliadas</strong><br>{len(rows)}</div>
    <div class="card"><strong>Erros</strong><br>{errors}</div>
    <div class="card"><strong>Alertas</strong><br>{warnings}</div>
  </div>
  <table>
    <thead><tr>{header}</tr></thead>
    <tbody>{table_rows}</tbody>
  </table>
</body>
</html>
""",
        encoding="utf-8",
    )


if __name__ == "__main__":
    raise SystemExit(main())

